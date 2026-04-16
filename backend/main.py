# ============================================================
#  main.py  –  FastAPI + MongoDB backend  v3
# ============================================================

import datetime, json, logging, os, uuid, copy
from contextlib import asynccontextmanager
from typing import Optional
import numpy as np
from motor.motor_asyncio import AsyncIOMotorDatabase

from auth import (
    LoginRequest, TokenResponse, MeResponse,
    login_user, refresh_token as _refresh_token, get_current_user,
)

import asyncio
from concurrent.futures import ThreadPoolExecutor
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Depends, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response, FileResponse
from pydantic import BaseModel

import config, data_loader
import solver as vrp_solver
import output_formatter, osrm_client
from database import (
    connect_db, close_db, get_db,
    DatasetDoc, StoreDoc, VehicleDoc, JobDoc, JobResultDoc, RunGroupDoc,
    bulk_insert_stores, bulk_insert_vehicles,
    save_matrix_bytes, load_matrix_bytes,
    save_excel_bytes, load_excel_bytes,
    db as mongo_db,
)

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-8s  %(message)s")
log = logging.getLogger(__name__)

_job_log_queues: dict[str, asyncio.Queue] = {}
_solve_executor  = ThreadPoolExecutor(max_workers=2, thread_name_prefix="vrp-solver")


class _JobLogHandler(logging.Handler):
    """
    Captures Python log records emitted during a solve run and routes
    them to a per-job asyncio.Queue so the browser WebSocket receives
    them in real time.
    """
    def __init__(self, job_id: str, loop: asyncio.AbstractEventLoop) -> None:
        super().__init__()
        self.job_id = job_id
        self.loop   = loop
        self.setFormatter(logging.Formatter("%(levelname)-8s %(name)s — %(message)s"))

    def emit(self, record: logging.LogRecord) -> None:
        q = _job_log_queues.get(self.job_id)
        if q is not None and self.loop.is_running():
            try:
                asyncio.run_coroutine_threadsafe(q.put(self.format(record)), self.loop)
            except Exception:
                pass


# ── App startup / shutdown via lifespan ──────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    await connect_db()   # connects Motor client + creates indexes
    yield
    await close_db()


app = FastAPI(title="VRP Route Optimization System", version="3.0.0",
              lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])


# ── POST /api/auth/login ──────────────────────────────────────
@app.post("/api/auth/login", response_model=TokenResponse)
async def auth_login(body: LoginRequest):
    """
    Accepts {username, password}.
    Returns {access_token, token_type, expires_in, username}.
    Token is valid for 30 minutes.
    """
    return login_user(body)
 

# ── POST /api/auth/refresh ────────────────────────────────────
@app.post("/api/auth/refresh", response_model=TokenResponse)
async def auth_refresh(username: str = Depends(get_current_user)):
    """
    Requires a valid Bearer token.
    Returns a fresh 30-minute token.
    The frontend calls this automatically when the token has < 5 min left.
    """
    return _refresh_token(username)
 

# ── POST /api/auth/logout ─────────────────────────────────────
@app.post("/api/auth/logout")
async def auth_logout(username: str = Depends(get_current_user)):
    """
    Server-side logout (stateless JWT — nothing to invalidate server-side,
    but the endpoint keeps the API consistent and lets you add a token
    blocklist later if needed).
    """
    log.info(f"User '{username}' logged out")
    return {"ok": True}
 

# ── GET /api/auth/me ──────────────────────────────────────────
@app.get("/api/auth/me")
async def auth_me(username: str = Depends(get_current_user)):
    """Returns the currently authenticated user. Used to restore sessions."""
    return {"username": username}

class _NumpySafeEncoder(json.JSONEncoder):
    """Converts numpy scalars to Python native types before JSON serialisation."""
    def default(self, obj):
        if isinstance(obj, np.integer):  return int(obj)
        if isinstance(obj, np.floating): return float(obj)
        if isinstance(obj, np.bool_):    return bool(obj)
        if isinstance(obj, np.ndarray):  return obj.tolist()
        return super().default(obj)


def _dumps(obj) -> str:
    return json.dumps(obj, cls=_NumpySafeEncoder)


# ════════════════════════════════════════════════════════════
#  Pydantic schemas
#  NOTE: dataset_id is now str (UUID) instead of int
# ════════════════════════════════════════════════════════════

class StoreUpdate(BaseModel):
    eng_name   : Optional[str]   = None
    mn_name    : Optional[str]   = None
    address    : Optional[str]   = None
    detail_addr: Optional[str]   = None
    lat        : Optional[float] = None
    lon        : Optional[float] = None
    open_s     : Optional[int]   = None
    close_s    : Optional[int]   = None
    dry_cbm    : Optional[float] = None
    dry_kg     : Optional[float] = None
    cold_cbm   : Optional[float] = None
    cold_kg    : Optional[float] = None

class StoreCreate(BaseModel):
    store_id   : str
    eng_name   : str   = ""
    mn_name    : str   = ""
    address    : str   = ""
    detail_addr: str   = ""
    lat        : float
    lon        : float
    open_s     : int   = 0
    close_s    : int   = 86399
    dry_cbm    : float = 0.0
    dry_kg     : float = 0.0
    cold_cbm   : float = 0.0
    cold_kg    : float = 0.0

class VehicleUpdate(BaseModel):
    description  : Optional[str]   = None
    depot        : Optional[str]   = None
    cap_kg       : Optional[float] = None
    cap_m3       : Optional[float] = None
    fuel_cost_km : Optional[float] = None
    vehicle_cost : Optional[float] = None
    labor_cost   : Optional[float] = None

class VehicleCreate(BaseModel):
    truck_id     : str
    description  : str   = ""
    depot        : str
    cap_kg       : float
    cap_m3       : float
    fuel_cost_km : float = 0.0
    vehicle_cost : float = 0.0
    labor_cost   : float = 0.0


# ════════════════════════════════════════════════════════════
#  Health
# ════════════════════════════════════════════════════════════

@app.get("/api/health")
async def health():
    osrm_ok = False
    try:
        import requests as rq
        r = rq.get(f"{config.OSRM_URL}/route/v1/driving/106.9,47.9;106.9,47.91",
                   timeout=3)
        osrm_ok = r.status_code == 200
    except Exception:
        pass
    return {"status": "ok", "osrm": "connected" if osrm_ok else "unreachable",
            "osrm_url": config.OSRM_URL, "version": "3.0.0"}


# ════════════════════════════════════════════════════════════
#  Dataset CRUD
# ════════════════════════════════════════════════════════════

@app.get("/api/datasets")
async def list_datasets(db: AsyncIOMotorDatabase = Depends(get_db)):
    cursor = db[DatasetDoc.COLLECTION].find().sort("created_at", -1)
    rows   = await cursor.to_list(None)
    out    = []
    for d in rows:
        store_count   = await db[StoreDoc.COLLECTION].count_documents({"dataset_id": d["_id"]})
        vehicle_count = await db[VehicleDoc.COLLECTION].count_documents({"dataset_id": d["_id"]})
        out.append({
            "id"           : d["_id"],
            "name"         : d["name"],
            "created_at"   : d["created_at"].isoformat(),
            "store_count"  : store_count,
            "vehicle_count": vehicle_count,
            "has_matrix"   : d.get("matrix_file_id") is not None,
        })
    return out


@app.post("/api/datasets")
async def create_dataset(
    name        : str                  = Form(...),
    store_file  : UploadFile           = File(...),
    matrix_file : Optional[UploadFile] = File(None),
    db          : AsyncIOMotorDatabase = Depends(get_db),
):
    store_bytes  = await store_file.read()
    matrix_bytes = await matrix_file.read() if matrix_file else None

    try:
        stores_list, import_warnings = data_loader.load_stores(store_bytes, "summer")
        vehicles_list = data_loader.load_vehicles(store_bytes)
    except Exception as e:
        raise HTTPException(422, f"Parse error: {e}")

    dataset_id  = str(uuid.uuid4())
    ds_doc      = DatasetDoc.make(name)
    ds_doc["_id"] = dataset_id
    await db[DatasetDoc.COLLECTION].insert_one(ds_doc)

    # Store binary matrix in GridFS
    if matrix_bytes:
        await save_matrix_bytes(dataset_id, matrix_bytes)

    await bulk_insert_stores(dataset_id, stores_list)
    await bulk_insert_vehicles(dataset_id, vehicles_list)

    return {"id": dataset_id, "name": name,
            "store_count": len(stores_list),
            "vehicle_count": len(vehicles_list)}


@app.delete("/api/datasets/{dataset_id}")
async def delete_dataset(dataset_id: str,
                         db: AsyncIOMotorDatabase = Depends(get_db)):
    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")

    # Unlink jobs & run_groups (preserve history)
    await db[JobDoc.COLLECTION].update_many(
        {"dataset_id": dataset_id}, {"$set": {"dataset_id": None}})
    await db[RunGroupDoc.COLLECTION].update_many(
        {"dataset_id": dataset_id}, {"$set": {"dataset_id": None}})

    # Delete stores + vehicles
    await db[StoreDoc.COLLECTION].delete_many({"dataset_id": dataset_id})
    await db[VehicleDoc.COLLECTION].delete_many({"dataset_id": dataset_id})

    # Delete dataset document
    await db[DatasetDoc.COLLECTION].delete_one({"_id": dataset_id})
    return {"ok": True}


@app.post("/api/datasets/{dataset_id}/matrix")
async def upload_matrix_to_dataset(
    dataset_id  : str,
    matrix_file : UploadFile           = File(...),
    db          : AsyncIOMotorDatabase = Depends(get_db),
):
    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")
    matrix_bytes = await matrix_file.read()
    await save_matrix_bytes(dataset_id, matrix_bytes)
    return {"ok": True, "dataset_id": dataset_id}


# ════════════════════════════════════════════════════════════
#  Dataset Export
# ════════════════════════════════════════════════════════════

@app.get("/api/datasets/{dataset_id}/export")
async def export_dataset(dataset_id: str,
                         db: AsyncIOMotorDatabase = Depends(get_db)):
    import pandas as pd
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")

    stores   = await db[StoreDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
    vehicles = await db[VehicleDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)

    def fmt_time(s: int) -> str:
        return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:00"

    # Create workbook with multi-level headers for seasonal data
    wb = Workbook()
    ws_stores = wb.active
    ws_stores.title = config.STORE_SHEET

    # Basic columns
    basic_cols = [
        config.COL_STORE_ID, config.COL_USE_YN, config.COL_ENG_NAME, config.COL_MN_NAME,
        config.COL_ADDR, config.COL_DTL_ADDR, "Location opening date", "Location closing date",
        config.COL_LAT, config.COL_LON, config.COL_OPEN, config.COL_CLOSE
    ]

    # Seasonal columns with multi-level headers
    seasons = ["Summer Avarage Order", "Autumn Avarage Order", "Winter Avarage Order", "Spring Avarage Order"]
    metrics = ["CBM (DRY DC)", "Weight (DRY DC)", "CBM (COLD DC)", "Weight (COLD DC)"]

    # Write multi-level headers
    col = 1
    for basic_col in basic_cols:
        ws_stores.cell(row=1, column=col, value=basic_col)
        ws_stores.cell(row=2, column=col, value="")
        col += 1

    for season in seasons:
        for metric in metrics:
            ws_stores.cell(row=1, column=col, value=season)
            ws_stores.cell(row=2, column=col, value=metric)
            col += 1

    # Style headers
    for row in [1, 2]:
        for c in range(1, col):
            cell = ws_stores.cell(row=row, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Write store data
    for row_idx, s in enumerate(stores, start=3):
        col = 1
        # Basic columns
        ws_stores.cell(row=row_idx, column=col, value=s["store_id"]); col += 1
        ws_stores.cell(row=row_idx, column=col, value="open"); col += 1
        ws_stores.cell(row=row_idx, column=col, value=s.get("eng_name", "")); col += 1
        ws_stores.cell(row=row_idx, column=col, value=s.get("mn_name", "")); col += 1
        ws_stores.cell(row=row_idx, column=col, value=s.get("address", "")); col += 1
        ws_stores.cell(row=row_idx, column=col, value=s.get("detail_addr", "")); col += 1
        ws_stores.cell(row=row_idx, column=col, value=""); col += 1  # Location opening date
        ws_stores.cell(row=row_idx, column=col, value=""); col += 1  # Location closing date
        ws_stores.cell(row=row_idx, column=col, value=s["lat"]); col += 1
        ws_stores.cell(row=row_idx, column=col, value=s["lon"]); col += 1
        ws_stores.cell(row=row_idx, column=col, value=fmt_time(s["open_s"])); col += 1
        ws_stores.cell(row=row_idx, column=col, value=fmt_time(s["close_s"])); col += 1

        # Seasonal data
        seasonal = s.get("seasonal_data", {})
        season_keys = ["summer", "autumn", "winter", "spring"]
        for s_key in season_keys:
            data = seasonal.get(s_key, {})
            ws_stores.cell(row=row_idx, column=col, value=data.get("dry_cbm", 0)); col += 1
            ws_stores.cell(row=row_idx, column=col, value=data.get("dry_kg", 0)); col += 1
            ws_stores.cell(row=row_idx, column=col, value=data.get("cold_cbm", 0)); col += 1
            ws_stores.cell(row=row_idx, column=col, value=data.get("cold_kg", 0)); col += 1

    # Create vehicles sheet
    ws_vehicles = wb.create_sheet(title=config.VEHICLE_SHEET)
    vehicles_data = [{
        config.COL_DEPOT       : v["depot"],
        config.COL_TRUCK_ID    : v["truck_id"],
        config.COL_DESCRIPTION : v.get("description", ""),
        config.COL_CAP_KG      : v["cap_kg"],
        config.COL_CAP_M3      : v["cap_m3"],
        config.COL_FUEL_COST   : v["fuel_cost_km"],
        config.COL_VEHICLE_COST: v["vehicle_cost"],
        config.COL_LABOR_COST  : v["labor_cost"],
    } for v in vehicles]

    for r_idx, row_data in enumerate(vehicles_data, start=1):
        for c_idx, (key, value) in enumerate(row_data.items(), start=1):
            ws_vehicles.cell(row=r_idx, column=c_idx, value=key if r_idx == 1 else value)
            if r_idx == 1:
                ws_vehicles.cell(row=r_idx, column=c_idx).font = Font(bold=True)

    # Add matrix sheets if available
    matrix_bytes = await load_matrix_bytes(dataset_id)
    if matrix_bytes:
        try:
            dist_df, dur_df = data_loader.load_matrix(matrix_bytes)
            # Duration sheet
            ws_duration = wb.create_sheet(title=config.DURATION_SHEET)
            for r_idx, (idx, row) in enumerate(dur_df.iterrows(), start=1):
                for c_idx, val in enumerate([idx] + row.tolist(), start=1):
                    ws_duration.cell(row=r_idx, column=c_idx, value=val)
            # Distance sheet
            ws_distance = wb.create_sheet(title=config.DISTANCE_SHEET)
            for r_idx, (idx, row) in enumerate(dist_df.iterrows(), start=1):
                for c_idx, val in enumerate([idx] + row.tolist(), start=1):
                    ws_distance.cell(row=r_idx, column=c_idx, value=val)
        except Exception as me:
            log.warning(f"Could not include matrix in export: {me}")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ds["name"].replace(" ", "_").replace("/", "-")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_dataset.xlsx"'},
    )


# ════════════════════════════════════════════════════════════
#  Store CRUD
# ════════════════════════════════════════════════════════════

@app.get("/api/datasets/{dataset_id}/stores")
async def list_stores(dataset_id: str,
                      db: AsyncIOMotorDatabase = Depends(get_db)):
    stores = await db[StoreDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
    # Convert MongoDB documents to JSON-serializable format
    result = []
    for s in stores:
        store_dict = dict(s)  # Convert to regular dict
        store_dict.pop("_id", None)  # Remove ObjectId field
        store_dict["id"] = str(s["_id"])  # Add string version
        result.append(store_dict)
    return result


@app.post("/api/datasets/{dataset_id}/stores")
async def add_store(dataset_id: str, body: StoreCreate,
                    db: AsyncIOMotorDatabase = Depends(get_db)):
    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")

    def norm(x):
        try: return str(int(str(x).strip()))
        except: return str(x).strip()

    doc = {
        "_id"        : str(uuid.uuid4()),
        "dataset_id" : dataset_id,
        "store_id"   : body.store_id,  # Preserve original format
        "node_id"    : norm(body.store_id),  # Normalized for matrix lookups
        "eng_name"   : body.eng_name,
        "mn_name"    : body.mn_name,
        "address"    : body.address,
        "detail_addr": body.detail_addr,
        "lat"        : body.lat,
        "lon"        : body.lon,
        "open_s"     : body.open_s,
        "close_s"    : body.close_s,
        "dry_cbm"    : body.dry_cbm,
        "dry_kg"     : body.dry_kg,
        "cold_cbm"   : body.cold_cbm,
        "cold_kg"    : body.cold_kg,
        "has_dry"    : body.dry_kg > 0 or body.dry_cbm > 0,
        "has_cold"   : body.cold_kg > 0 or body.cold_cbm > 0,
        "seasonal_data": {},  # Empty for manually added stores
    }
    await db[StoreDoc.COLLECTION].insert_one(doc)
    # Convert MongoDB document to JSON-serializable format
    result_dict = dict(doc)
    result_dict.pop("_id", None)
    result_dict["id"] = doc["_id"]
    return result_dict


@app.put("/api/datasets/{dataset_id}/stores/{sid}")
async def update_store(dataset_id: str, sid: str, body: StoreUpdate,
                       db: AsyncIOMotorDatabase = Depends(get_db)):
    s = await db[StoreDoc.COLLECTION].find_one({"_id": sid, "dataset_id": dataset_id})
    if not s:
        raise HTTPException(404, "Store not found")

    updates = body.model_dump(exclude_none=True)
    # Recompute flags if demand fields changed
    dry_kg   = updates.get("dry_kg",   s["dry_kg"])
    dry_cbm  = updates.get("dry_cbm",  s["dry_cbm"])
    cold_kg  = updates.get("cold_kg",  s["cold_kg"])
    cold_cbm = updates.get("cold_cbm", s["cold_cbm"])
    updates["has_dry"]  = dry_kg > 0  or dry_cbm > 0
    updates["has_cold"] = cold_kg > 0 or cold_cbm > 0

    await db[StoreDoc.COLLECTION].update_one({"_id": sid}, {"$set": updates})
    updated = await db[StoreDoc.COLLECTION].find_one({"_id": sid})
    # Convert MongoDB document to JSON-serializable format
    result_dict = dict(updated)
    result_dict.pop("_id", None)
    result_dict["id"] = updated["_id"]
    return result_dict


@app.delete("/api/datasets/{dataset_id}/stores/{sid}")
async def delete_store(dataset_id: str, sid: str,
                       db: AsyncIOMotorDatabase = Depends(get_db)):
    r = await db[StoreDoc.COLLECTION].delete_one({"_id": sid, "dataset_id": dataset_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Store not found")
    return {"ok": True}


@app.delete("/api/datasets/{dataset_id}/stores")
async def clear_stores(dataset_id: str,
                       db: AsyncIOMotorDatabase = Depends(get_db)):
    await db[StoreDoc.COLLECTION].delete_many({"dataset_id": dataset_id})
    return {"ok": True}


# ════════════════════════════════════════════════════════════
#  Vehicle CRUD
# ════════════════════════════════════════════════════════════

@app.get("/api/datasets/{dataset_id}/vehicles")
async def list_vehicles(dataset_id: str,
                        db: AsyncIOMotorDatabase = Depends(get_db)):
    vehicles = await db[VehicleDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
    # Convert MongoDB documents to JSON-serializable format
    result = []
    for v in vehicles:
        vehicle_dict = dict(v)  # Convert to regular dict
        vehicle_dict.pop("_id", None)  # Remove ObjectId field
        vehicle_dict["id"] = str(v["_id"])  # Add string version
        result.append(vehicle_dict)
    return result


@app.post("/api/datasets/{dataset_id}/vehicles")
async def add_vehicle(dataset_id: str, body: VehicleCreate,
                      db: AsyncIOMotorDatabase = Depends(get_db)):
    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")

    fleet = config.DEPOT_VEHICLE_MAP.get(body.depot, "DRY")
    doc = {
        "_id"         : str(uuid.uuid4()),
        "dataset_id"  : dataset_id,
        "truck_id"    : body.truck_id,
        "description" : body.description,
        "depot"       : body.depot,
        "fleet"       : fleet,
        "cap_kg"      : body.cap_kg,
        "cap_m3"      : body.cap_m3,
        "fuel_cost_km": body.fuel_cost_km,
        "vehicle_cost": body.vehicle_cost,
        "labor_cost"  : body.labor_cost,
    }
    await db[VehicleDoc.COLLECTION].insert_one(doc)
    # Convert MongoDB document to JSON-serializable format
    result_dict = dict(doc)
    result_dict.pop("_id", None)
    result_dict["id"] = doc["_id"]
    return result_dict


@app.put("/api/datasets/{dataset_id}/vehicles/{vid}")
async def update_vehicle(dataset_id: str, vid: str, body: VehicleUpdate,
                         db: AsyncIOMotorDatabase = Depends(get_db)):
    v = await db[VehicleDoc.COLLECTION].find_one({"_id": vid, "dataset_id": dataset_id})
    if not v:
        raise HTTPException(404, "Vehicle not found")

    updates = body.model_dump(exclude_none=True)
    # Recompute fleet if depot changed
    depot = updates.get("depot", v.get("depot"))
    if depot:
        updates["fleet"] = config.DEPOT_VEHICLE_MAP.get(depot, "DRY")

    await db[VehicleDoc.COLLECTION].update_one({"_id": vid}, {"$set": updates})
    updated = await db[VehicleDoc.COLLECTION].find_one({"_id": vid})
    # Convert MongoDB document to JSON-serializable format
    result_dict = dict(updated)
    result_dict.pop("_id", None)
    result_dict["id"] = updated["_id"]
    return result_dict


@app.delete("/api/datasets/{dataset_id}/vehicles/{vid}")
async def delete_vehicle(dataset_id: str, vid: str,
                         db: AsyncIOMotorDatabase = Depends(get_db)):
    r = await db[VehicleDoc.COLLECTION].delete_one({"_id": vid, "dataset_id": dataset_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Vehicle not found")
    return {"ok": True}


@app.delete("/api/datasets/{dataset_id}/vehicles")
async def clear_vehicles(dataset_id: str,
                         db: AsyncIOMotorDatabase = Depends(get_db)):
    await db[VehicleDoc.COLLECTION].delete_many({"dataset_id": dataset_id})
    return {"ok": True}


# ════════════════════════════════════════════════════════════
#  Optimize
# ════════════════════════════════════════════════════════════

@app.post("/api/optimize")
async def optimize(
    dataset_id        : Optional[str]          = Form(None),
    store_file        : Optional[UploadFile]   = File(None),
    matrix_file       : Optional[UploadFile]   = File(None),
    mode              : str                    = Form("cheapest"),
    max_trips         : int                    = Form(3),
    solver_time       : int                    = Form(120),
    rural_solver_time : Optional[int]          = Form(None),
    group_id          : Optional[str]          = Form(None),
    version_name      : Optional[str]          = Form(None),
    season            : str                    = Form("summer"),
    max_weight_fill   : float                  = Form(1.0),
    max_volume_fill   : float                  = Form(1.0),
    custom_config     : Optional[str]          = Form(None),
    db                : AsyncIOMotorDatabase   = Depends(get_db),
):
    """
    Start an optimization run.

    Returns {job_id, status:"running"} immediately.
    The solve runs as a background asyncio task.
    Stream real-time logs via  WS /ws/logs/{job_id}
    Poll for completion via    GET /api/jobs/{job_id}
    """
    if mode not in ("fastest", "shortest", "cheapest", "balanced", "geographic"):
        raise HTTPException(400, f"Invalid mode '{mode}'")
    if not (0.0 <= max_weight_fill <= 1.5):
        raise HTTPException(400, "max_weight_fill must be between 0.0 and 1.5")
    if not (0.0 <= max_volume_fill <= 1.5):
        raise HTTPException(400, "max_volume_fill must be between 0.0 and 1.5")

    # ── Resolve data (fast — only I/O) ──────────────────────
    if dataset_id:
        ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
        if not ds:
            raise HTTPException(404, "Dataset not found")
        store_docs    = await db[StoreDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
        vehicle_docs  = await db[VehicleDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
        stores_list   = [StoreDoc.to_solver_dict(s) for s in store_docs]
        vehicles_list = [VehicleDoc.to_solver_dict(v) for v in vehicle_docs]
        matrix_bytes  = await load_matrix_bytes(dataset_id)
        if not matrix_bytes:
            raise HTTPException(422, "No matrix for this dataset — upload it first")
    elif store_file and matrix_file:
        sb            = await store_file.read()
        matrix_bytes  = await matrix_file.read()
        stores_list, _ = data_loader.load_stores(sb, season)  # Ignore warnings for optimization
        vehicles_list = data_loader.load_vehicles(sb)
        dataset_id    = None
    else:
        raise HTTPException(422, "Provide dataset_id OR both store_file+matrix_file")

    # ── Create job record ────────────────────────────────────
    job_id = str(uuid.uuid4())
    if not version_name and group_id:
        count = await db[JobDoc.COLLECTION].count_documents({"group_id": group_id})
        version_name = f"Auto v{count + 1}"
    elif not version_name:
        version_name = "Auto v1"

    job_doc = JobDoc.make(
        job_id, dataset_id=dataset_id, group_id=group_id,
        version_name=version_name, mode=mode,
        max_trips=max_trips, solver_time=solver_time,
        rural_solver_time=rural_solver_time,
        season=season,
    )
    job_doc["status"] = "running"
    await db[JobDoc.COLLECTION].insert_one(job_doc)

    # ── Parse custom config if provided ─────────────────────
    custom_config_dict = None
    if custom_config:
        try:
            custom_config_dict = json.loads(custom_config)
            log.info(f"Job {job_id[:8]}: Using custom config overrides: {list(custom_config_dict.keys())}")
        except json.JSONDecodeError:
            log.warning(f"Job {job_id[:8]}: Invalid custom_config JSON, ignoring")

    # ── Fire-and-forget background task ─────────────────────
    asyncio.create_task(_solve_background(
        job_id, stores_list, vehicles_list, matrix_bytes,
        mode, max_trips, solver_time, max_weight_fill, max_volume_fill,
        rural_solver_time, season, db, custom_config_dict,
    ))

    log.info(f"Job {job_id[:8]} queued — mode={mode}, stores={len(stores_list)}, solver={solver_time}s")

    # Return immediately — frontend polls GET /api/jobs/{job_id} for completion
    return {"job_id": job_id, "status": "running"}

# ════════════════════════════════════════════════════════════
#  Manual Job Creation
# ════════════════════════════════════════════════════════════

class ManualJobCreate(BaseModel):
    title     : str
    routes    : list[dict]
    is_manual : bool
    dataset_id: str          # ← str now


@app.post("/api/jobs/manual")
async def create_manual_job(body: ManualJobCreate,
                            db: AsyncIOMotorDatabase = Depends(get_db)):
    ds = await db[DatasetDoc.COLLECTION].find_one({"_id": body.dataset_id})
    if not ds:
        raise HTTPException(404, "Dataset not found")

    matrix_bytes = await load_matrix_bytes(body.dataset_id)
    if not matrix_bytes:
        raise HTTPException(422, "Dataset must have a distance matrix for manual routing")

    for route in body.routes:
        if not route.get("vehicle_id"):
            raise HTTPException(400, "All routes must have a vehicle_id")
        if not route.get("stops") or len(route["stops"]) == 0:
            raise HTTPException(400, "All routes must have at least one stop")

    try:
        dist_df, dur_df = data_loader.load_matrix(matrix_bytes)
    except Exception as e:
        raise HTTPException(422, f"Matrix error: {e}")

    store_docs   = await db[StoreDoc.COLLECTION].find({"dataset_id": body.dataset_id}).to_list(None)
    vehicle_docs = await db[VehicleDoc.COLLECTION].find({"dataset_id": body.dataset_id}).to_list(None)
    stores_dict  = {s["store_id"]: s for s in store_docs}
    vehicles_dict = {v["truck_id"]: v for v in vehicle_docs}

    job_id  = str(uuid.uuid4())
    job_doc = JobDoc.make(
        job_id, dataset_id=body.dataset_id,
        version_name=body.title, is_manual=True,
        mode="manual", max_trips=1, solver_time=0,
    )
    job_doc["status"]       = "done"
    job_doc["completed_at"] = datetime.datetime.utcnow()
    await db[JobDoc.COLLECTION].insert_one(job_doc)

    route_summary : list = []
    stop_details  : list = []
    wps_map       : dict = {}
    served_dry    : set  = set()
    served_cold   : set  = set()
    total_cost     = 0.0
    total_dist_km  = 0.0
    trip_counter  : dict = {}

    COLORS = [
        "#5B7CFA","#22D3EE","#34D399","#A78BFA","#F472B6",
        "#38BDF8","#4ADE80","#818CF8","#FB7185","#2DD4BF",
        "#F97316","#EAB308","#84CC16","#DC2626","#D97706",
    ]

    def fmt_wall(s: float) -> str:
        h = int(s // 3600) % 24
        m = int((s % 3600) // 60)
        return f"{h:02d}:{m:02d}"

    def matrix_lookup(from_id: str, to_id: str):
        if from_id in dist_df.index and to_id in dist_df.columns:
            dist = float(dist_df.loc[from_id, to_id])
            dur  = float(dur_df.loc[from_id, to_id]) * 60.0
            return dist, dur
        return 0.0, 0.0

    for route_idx, route in enumerate(body.routes):
        vehicle_id = route["vehicle_id"]
        stop_ids   = [s.strip() for s in route.get("stops", []) if str(s).strip()]
        route_name = route.get("route_name", f"Route {route_idx + 1}")
        truck_number = route.get("truck_number")
        contractor = route.get("contractor")

        vehicle = vehicles_dict.get(vehicle_id)
        if not vehicle:
            raise HTTPException(400, f"Vehicle {vehicle_id} not found in dataset")

        fleet      = vehicle["fleet"]
        sched      = config.FLEET_SCHEDULE.get(fleet, {"start_hour": 8, "max_horizon_hour": 20})
        start_wall = float(sched["start_hour"] * 3600)

        trip_counter[vehicle_id] = trip_counter.get(vehicle_id, 0) + 1
        trip_num = trip_counter[vehicle_id]

        depot_name = "Dry DC" if fleet == "DRY" else "Cold DC"
        depot_cfg  = config.DEPOTS[depot_name]

        valid_stop_ids = [sid for sid in stop_ids if sid in stores_dict]
        node_seq = ([depot_name]
                    + [stores_dict[sid]["node_id"] for sid in valid_stop_ids]
                    + [depot_name])

        leg_dists, leg_durs = [], []
        for i in range(len(node_seq) - 1):
            dm, ds_ = matrix_lookup(node_seq[i], node_seq[i + 1])
            leg_dists.append(dm); leg_durs.append(ds_)

        total_route_dist = sum(leg_dists)
        current_wall = start_wall
        route_load_kg = route_load_m3 = 0.0
        route_stops: list = []

        for si, sid in enumerate(valid_stop_ids):
            store    = stores_dict[sid]
            arr_wall = current_wall + (leg_durs[si] if si < len(leg_durs) else 0.0)

            demand_kg = float(store["dry_kg"]  if fleet == "DRY" else store["cold_kg"])
            demand_m3 = float(store["dry_cbm"] if fleet == "DRY" else store["cold_cbm"])
            route_load_kg += demand_kg
            route_load_m3 += demand_m3

            dep_wall     = arr_wall + config.SERVICE_TIME_SECONDS
            current_wall = dep_wall
            day_num      = 1 + int(arr_wall // 86400)

            route_stops.append({
                "fleet"       : fleet,
                "truck_id"    : vehicle_id,
                "trip_number" : trip_num,
                "stop_order"  : si + 1,
                "store_id"    : sid,
                "eng_name"    : store.get("eng_name", ""),
                "mn_name"     : store.get("mn_name", ""),
                "address"     : store.get("address", ""),
                "detail_addr" : store.get("detail_addr", ""),
                "lat"         : store["lat"],
                "lon"         : store["lon"],
                "arrival"     : fmt_wall(arr_wall),
                "departure"   : fmt_wall(dep_wall),
                "delivery_day": "Same day" if day_num <= 1 else f"Day {day_num}",
                "is_rural"    : False,
                "demand_kg"   : round(demand_kg, 2),
                "demand_m3"   : round(demand_m3, 3),
                "truck_num"   : truck_number,
                "contractor"  : contractor or vehicle.get("contractor"),
            })

            if fleet == "DRY": served_dry.add(sid)
            else:              served_cold.add(sid)

        return_dur  = leg_durs[-1] if leg_durs else 0.0
        return_wall = current_wall + return_dur

        dist_km    = total_route_dist / 1000.0
        fuel_cost  = dist_km * float(vehicle["fuel_cost_km"])
        route_cost = fuel_cost + float(vehicle["vehicle_cost"]) + float(vehicle["labor_cost"])
        total_cost    += route_cost
        total_dist_km += dist_km
        cap_kg = float(vehicle["cap_kg"])
        cap_m3 = float(vehicle["cap_m3"])

        route_summary.append({
            "fleet"       : fleet,
            "truck_id"    : vehicle_id,
            "trip_number" : trip_num,
            "route_type"  : "manual",
            "stops"       : len(valid_stop_ids),
            "distance_km" : round(dist_km, 1),
            "duration_min": round((return_wall - start_wall) / 60.0, 1),
            "load_kg"     : round(route_load_kg, 2),
            "cap_kg"      : cap_kg,
            "util_kg_pct" : round(route_load_kg / cap_kg * 100, 1) if cap_kg else 0.0,
            "load_m3"     : round(route_load_m3, 3),
            "cap_m3"      : cap_m3,
            "util_m3_pct" : round(route_load_m3 / cap_m3 * 100, 1) if cap_m3 else 0.0,
            "cost_fuel"   : round(fuel_cost, 0),
            "cost_fixed"  : float(vehicle["vehicle_cost"]),
            "cost_labor"  : float(vehicle["labor_cost"]),
            "cost_total"  : round(route_cost, 0),
            "departs_at"  : fmt_wall(start_wall),
            "truck_num"   : truck_number,
            "contractor"  : contractor or vehicle.get("contractor"),
            "returns_at"  : fmt_wall(return_wall),
            "is_overnight": return_wall >= 86400,
            "man_hours"   : round((return_wall - start_wall) / 3600.0, 2),
        })
        stop_details.extend(route_stops)

        osrm_wps = [(depot_cfg["lat"], depot_cfg["lon"])]
        for sid in valid_stop_ids:
            st = stores_dict[sid]
            osrm_wps.append((st["lat"], st["lon"]))
        osrm_wps.append((depot_cfg["lat"], depot_cfg["lon"]))
        wps_map[f"{vehicle_id}_T{trip_num}"] = osrm_wps

    # ── OSRM geometries ───────────────────────────────────────
    raw_geometries: dict = {}
    try:
        raw_geometries = osrm_client.get_route_geometries_batch(wps_map)
    except Exception as e:
        log.warning(f"OSRM geometry batch failed: {e}")

    # ── Map data ──────────────────────────────────────────────
    map_data: list = []
    color_counters = {"DRY": 0, "COLD": 8}

    for rs in route_summary:
        vid   = rs["truck_id"]
        tnum  = rs["trip_number"]
        fleet = rs["fleet"]
        rid   = f"{vid}_T{tnum}"
        dc_name = "Dry DC" if fleet == "DRY" else "Cold DC"
        dc_cfg  = config.DEPOTS[dc_name]

        color_idx = color_counters[fleet] % len(COLORS)
        color_counters[fleet] += 1
        color = COLORS[color_idx]

        route_stop_objs = sorted(
            [s for s in stop_details if s["truck_id"] == vid and s["trip_number"] == tnum],
            key=lambda x: x["stop_order"],
        )
        map_stops = [{
            "lat"        : s["lat"],
            "lon"        : s["lon"],
            "order"      : s["stop_order"],
            "store_id"   : s["store_id"],
            "name"       : s["eng_name"],
            "mn_name"    : s["mn_name"],
            "arrival"    : s["arrival"],
            "day_label"  : "" if s["delivery_day"] == "Same day" else s["delivery_day"],
            "is_rural"   : False,
            "is_next_day": s["delivery_day"] != "Same day",
            "demand_kg"  : s["demand_kg"],
            "demand_m3"  : s["demand_m3"],
        } for s in route_stop_objs]

        raw_geo = raw_geometries.get(rid)
        if raw_geo:
            polyline = [[pt[1], pt[0]] for pt in raw_geo]
        else:
            polyline = [[dc_cfg["lat"], dc_cfg["lon"]]]
            for ms in map_stops:
                polyline.append([ms["lat"], ms["lon"]])
            polyline.append([dc_cfg["lat"], dc_cfg["lon"]])

        map_data.append({
            "route_id"   : rid,
            "fleet"      : fleet,
            "truck_id"   : vid,
            "trip_number": tnum,
            "is_rural"   : False,
            "color"      : color,
            "line_style" : "solid",
            "stops"      : map_stops,
            "polyline"   : polyline,
            "depot_lat"  : dc_cfg["lat"],
            "depot_lon"  : dc_cfg["lon"],
            "sched_info" : f"Manual · Departs {rs['departs_at']} · Returns {rs['returns_at']}",
            "summary"    : {
                "distance_km" : rs["distance_km"],
                "duration_min": rs["duration_min"],
                "load_kg"     : rs["load_kg"],
                "load_m3"     : rs["load_m3"],
                "return_at"   : rs["returns_at"],
                "is_overnight": rs["is_overnight"],
            },
        })

    # ── Unserved ──────────────────────────────────────────────
    unserved: list = []
    for s in store_docs:
        if s.get("has_dry") and s["store_id"] not in served_dry:
            unserved.append({
                "fleet"    : "DRY",
                "store_id" : s["store_id"],
                "eng_name" : s.get("eng_name", ""),
                "mn_name"  : s.get("mn_name", ""),
                "address"  : s.get("address", ""),
                "lat"      : s["lat"],
                "lon"      : s["lon"],
                "demand_kg": round(float(s["dry_kg"]),  2),
                "demand_m3": round(float(s["dry_cbm"]), 3),
                "reason"   : "Not assigned to any route in manual plan.",
            })
        if s.get("has_cold") and s["store_id"] not in served_cold:
            unserved.append({
                "fleet"    : "COLD",
                "store_id" : s["store_id"],
                "eng_name" : s.get("eng_name", ""),
                "mn_name"  : s.get("mn_name", ""),
                "address"  : s.get("address", ""),
                "lat"      : s["lat"],
                "lon"      : s["lon"],
                "demand_kg": round(float(s["cold_kg"]),  2),
                "demand_m3": round(float(s["cold_cbm"]), 3),
                "reason"   : "Not assigned to any route in manual plan.",
            })

    total_served = sum(r["stops"] for r in route_summary)
    summary = {
        "mode"           : "manual",
        "total_stores"   : len(stores_dict),
        "total_served"   : total_served,
        "total_unserved" : len(unserved),
        "total_routes"   : len(route_summary),
        "total_dist_km"  : round(total_dist_km, 1),
        "total_cost"     : round(total_cost, 0),
        "total_man_hours": round(sum(r.get("man_hours", 0) for r in route_summary), 1),
        "warnings"       : [],
    }

    try:
        excel_bytes = output_formatter.export_to_excel(route_summary, stop_details, unserved)
    except Exception as e:
        log.warning(f"Excel generation failed for manual job {job_id}: {e}")
        excel_bytes = b""

    result_doc = JobResultDoc.make(job_id, summary, route_summary,
                                   stop_details, unserved, map_data)
    await db[JobResultDoc.COLLECTION].insert_one(result_doc)
    if excel_bytes:
        await save_excel_bytes(job_id, excel_bytes)

    log.info(f"Manual job {job_id[:8]} done — {total_served} served, "
             f"{len(unserved)} unserved, {len(route_summary)} routes")
    return JobDoc.to_dict(job_doc)

async def _solve_background(
    job_id          : str,
    stores_list     : list,
    vehicles_list   : list,
    matrix_bytes    : bytes,
    mode            : str,
    max_trips       : int,
    solver_time     : int,
    max_weight_fill : float,
    max_volume_fill : float,
    rural_solver_time: Optional[int],
    season          : str,
    db,
    custom_config   : Optional[dict] = None,
):
    from solver import SolverConfig   # local import keeps top-level clean

    loop = asyncio.get_event_loop()
    q    = asyncio.Queue()
    _job_log_queues[job_id] = q

    handler  = _JobLogHandler(job_id, loop)
    root_log = logging.getLogger()
    root_log.addHandler(handler)

    async def _emit(msg: str):
        await q.put(msg)

    try:
        await _emit(f"INFO     main — ⚡ Job {job_id[:8]} starting …")

        # ── Build SolverConfig (no global mutation) ──────────────
        cfg = SolverConfig(
            mode                  = mode,
            max_trips             = max_trips,
            solver_time_s         = solver_time,
            rural_solver_time     = rural_solver_time,
            max_weight_fill       = max_weight_fill,
            max_volume_fill       = max_volume_fill,
            reload_time_s         = config.RELOAD_TIME_SECONDS,
            service_time_base_s   = config.SERVICE_TIME_SECONDS,
            service_time_per_kg_s = 0.0,
            penalty_unserved      = config.PENALTY_UNSERVED,
            vehicle_fixed_cost    = config.VEHICLE_FIXED_COST,
            m3_scale              = config.M3_SCALE,
            far_threshold_km      = config.FAR_THRESHOLD_KM,
        )

        # ── Apply custom config overrides if provided ───────────
        if custom_config:
            # Handle FLEET_SCHEDULE overrides
            fleet_schedule_overrides = {}
            for key in ["DRY_START_HOUR", "DRY_MAX_HORIZON_HOUR", "COLD_START_HOUR", "COLD_MAX_HORIZON_HOUR"]:
                if key in custom_config:
                    fleet_schedule_overrides[key] = custom_config[key]

            if fleet_schedule_overrides:
                # Reconstruct FLEET_SCHEDULE with overrides
                original_schedule = copy.deepcopy(config.FLEET_SCHEDULE)
                if "DRY_START_HOUR" in fleet_schedule_overrides:
                    original_schedule["DRY"]["start_hour"] = fleet_schedule_overrides["DRY_START_HOUR"]
                if "DRY_MAX_HORIZON_HOUR" in fleet_schedule_overrides:
                    original_schedule["DRY"]["max_horizon_hour"] = fleet_schedule_overrides["DRY_MAX_HORIZON_HOUR"]
                if "COLD_START_HOUR" in fleet_schedule_overrides:
                    original_schedule["COLD"]["start_hour"] = fleet_schedule_overrides["COLD_START_HOUR"]
                if "COLD_MAX_HORIZON_HOUR" in fleet_schedule_overrides:
                    original_schedule["COLD"]["max_horizon_hour"] = fleet_schedule_overrides["COLD_MAX_HORIZON_HOUR"]
                config.FLEET_SCHEDULE = original_schedule
                await _emit(f"INFO     main — Custom FLEET_SCHEDULE applied")

            # Handle other config overrides
            for key, value in custom_config.items():
                if key in fleet_schedule_overrides:
                    continue  # Already handled above
                if hasattr(cfg, key):
                    setattr(cfg, key, value)
                    await _emit(f"INFO     main — Custom config: {key} = {value}")
                elif hasattr(config, key):
                    setattr(config, key, value)
                    await _emit(f"INFO     main — Custom config (global): {key} = {value}")
                else:
                    await _emit(f"WARNING  main — Unknown config key: {key}")
 
        await _emit(f"INFO     main — mode={cfg.mode}, trips={cfg.max_trips}, "
                    f"budget={cfg.solver_time_s}s")
 
        # ── Load matrices ────────────────────────────────────────
        await _emit("INFO     main — Loading distance matrix …")
        dist_df, dur_df = await loop.run_in_executor(
            None, lambda: data_loader.load_matrix(matrix_bytes)
        )
        warnings = data_loader.validate_data(
            stores_list, vehicles_list, dist_df, dur_df
        )
        for w in warnings:
            await _emit(f"WARNING  main — {w}")
 
        dry_stores  = sum(1 for s in stores_list if s.get("has_dry"))
        cold_stores = sum(1 for s in stores_list if s.get("has_cold"))
        dry_v       = sum(1 for v in vehicles_list if v.get("fleet") == "DRY")
        cold_v      = sum(1 for v in vehicles_list if v.get("fleet") == "COLD")
        await _emit(
            f"INFO     main — DRY: {dry_stores} stores / {dry_v} trucks | "
            f"COLD: {cold_stores} stores / {cold_v} trucks"
        )
 
        # ── Solve ────────────────────────────────────────────────
        await _emit("INFO     main — Starting OR-Tools solver …")
        result = await loop.run_in_executor(
            _solve_executor,
            lambda: vrp_solver.solve(
                stores_list, vehicles_list, dist_df, dur_df, cfg=cfg, season=season
            ),
        )
 
        # ── OSRM geometries ──────────────────────────────────────
        await _emit("INFO     osrm — Fetching route polylines …")
        route_geometries: dict = {}
        try:
            wps_map: dict = {}
            for fleet, fr in result.items():
                for route in fr["routes"]:
                    dc  = (config.DEPOTS["Dry DC"] if fleet == "DRY"
                           else config.DEPOTS["Cold DC"])
                    wps = [(dc["lat"], dc["lon"])]
                    for stop in route["stops"]:
                        wps.append((stop["lat"], stop["lon"]))
                    wps.append((dc["lat"], dc["lon"]))
                    wps_map[route["virtual_id"]] = wps
 
            route_geometries = await loop.run_in_executor(
                None, lambda: osrm_client.get_route_geometries_batch(wps_map)
            )
            await _emit(
                f"INFO     osrm — {len(route_geometries)} route geometries fetched"
            )
        except Exception as e:
            await _emit(f"WARNING  osrm — Geometry skipped: {e}")
 
        # ── Format + persist ─────────────────────────────────────
        route_summary = output_formatter.build_route_summary(result)
        stop_details  = output_formatter.build_stop_details(result)
        unserved      = output_formatter.build_unserved(result, dist_df)
        map_data      = output_formatter.build_map_data(result, route_geometries)
 
        served     = sum(len(r["stops"]) for fr in result.values() for r in fr["routes"])
        unserved_n = sum(len(fr["unserved"]) for fr in result.values())
        total_cost = sum(r["cost_total"]  for r in route_summary)
        total_dist = sum(r["distance_km"] for r in route_summary)
        total_mh   = sum(r.get("man_hours", 0) for r in route_summary)
 
        summary = {
            "mode"           : mode,
            "total_stores"   : len(stores_list),
            "total_served"   : served,
            "total_unserved" : unserved_n,
            "total_routes"   : len(route_summary),
            "total_dist_km"  : round(total_dist, 1),
            "total_cost"     : round(total_cost, 0),
            "total_man_hours": round(total_mh, 1),
            "warnings"       : warnings,
        }
 
        await _emit(
            f"INFO     main — ✅ Done: {served} served · {unserved_n} unserved · "
            f"{len(route_summary)} routes · ₮{round(total_cost):,}"
        )
 
        await _emit("INFO     main — Saving results …")
        excel_bytes = await loop.run_in_executor(
            None,
            lambda: output_formatter.export_to_excel(
                route_summary, stop_details, unserved
            ),
        )
        result_doc = JobResultDoc.make(
            job_id, summary, route_summary, stop_details, unserved, map_data
        )
        await db[JobResultDoc.COLLECTION].insert_one(result_doc)
        await save_excel_bytes(job_id, excel_bytes)
 
        await db[JobDoc.COLLECTION].update_one(
            {"_id": job_id},
            {"$set": {
                "status"      : "done",
                "completed_at": datetime.datetime.utcnow(),
            }},
        )
        log.info(f"Job {job_id[:8]} completed — {served} served, mode={mode}")
 
    except Exception as exc:
        log.exception(f"Job {job_id[:8]} failed: {exc}")
        await _emit(f"ERROR    main — Solver crashed: {exc}")
        try:
            await db[JobDoc.COLLECTION].update_one(
                {"_id": job_id},
                {"$set": {"status": "error", "error_msg": str(exc)}},
            )
        except Exception:
            pass
 
    finally:
        root_log.removeHandler(handler)
        await q.put("__DONE__")
 

# ════════════════════════════════════════════════════════════
#  Jobs
# ════════════════════════════════════════════════════════════

@app.get("/api/jobs")
async def list_jobs(limit: int = 30, db: AsyncIOMotorDatabase = Depends(get_db)):
    cursor = db[JobDoc.COLLECTION].find().sort("created_at", -1).limit(limit)
    jobs   = await cursor.to_list(None)
    out    = []
    for j in jobs:
        d = JobDoc.to_dict(j)
        res = await db[JobResultDoc.COLLECTION].find_one({"_id": j["_id"]})
        if res:
            s = JobResultDoc.get_summary(res)
            d["total_served"]    = s.get("total_served")    if s else None
            d["total_unserved"]  = s.get("total_unserved")  if s else None
            d["total_routes"]    = s.get("total_routes")    if s else None
            d["total_cost"]      = s.get("total_cost")      if s else None
            d["total_man_hours"] = s.get("total_man_hours") if s else None
        out.append(d)
    return out


@app.get("/api/jobs/{job_id}")
async def get_job_result(job_id: str, db: AsyncIOMotorDatabase = Depends(get_db)):
    job = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    info = JobDoc.to_dict(job)
    res  = await db[JobResultDoc.COLLECTION].find_one({"_id": job_id})
    if res:
        info.update({
            "summary"      : JobResultDoc.get_summary(res),
            "route_summary": JobResultDoc.get_routes(res),
            "stop_details" : JobResultDoc.get_stops(res),
            "unserved"     : JobResultDoc.get_unserved(res),
            "map_data"     : JobResultDoc.get_map_data(res),
        })
    return info


@app.delete("/api/jobs/{job_id}")
async def delete_job(job_id: str, db: AsyncIOMotorDatabase = Depends(get_db)):
    job = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    await db[JobDoc.COLLECTION].delete_one({"_id": job_id})
    await db[JobResultDoc.COLLECTION].delete_one({"_id": job_id})
    return {"ok": True}


# ════════════════════════════════════════════════════════════
#  Excel export
# ════════════════════════════════════════════════════════════

@app.get("/api/export/{job_id}")
async def export_excel(job_id: str, db: AsyncIOMotorDatabase = Depends(get_db)):
    job = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    excel_bytes = await load_excel_bytes(job_id)
    if not excel_bytes:
        raise HTTPException(404, "Job result not found. Run optimization first.")
    return Response(
        content    = excel_bytes,
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition":
                      f'attachment; filename="vrp_{job_id[:8]}_{job["mode"]}.xlsx"'},
    )


# ════════════════════════════════════════════════════════════
#  Build Matrix
# ════════════════════════════════════════════════════════════

@app.post("/api/build-matrix")
async def build_matrix_endpoint(
    dataset_id      : Optional[str]            = Form(None),
    store_file      : Optional[UploadFile]     = File(None),
    matrix_file     : Optional[UploadFile]     = File(None),
    save_to_dataset : bool                     = Form(True),
    db              : AsyncIOMotorDatabase     = Depends(get_db),
):
    if dataset_id:
        ds = await db[DatasetDoc.COLLECTION].find_one({"_id": dataset_id})
        if not ds:
            raise HTTPException(404, "Dataset not found")
        store_docs  = await db[StoreDoc.COLLECTION].find({"dataset_id": dataset_id}).to_list(None)
        stores_list = [StoreDoc.to_solver_dict(s) for s in store_docs]
    elif store_file:
        stores_list, _ = data_loader.load_stores(await store_file.read(), "summer")  # Ignore warnings for optimization
        dataset_id  = None
    else:
        raise HTTPException(422, "Provide dataset_id or store_file")

    if matrix_file:
        matrix_bytes = await matrix_file.read()
        try:
            import pandas as pd, io as _io
            df = pd.read_excel(_io.BytesIO(matrix_bytes), sheet_name=None)
            if config.DURATION_SHEET not in df or config.DISTANCE_SHEET not in df:
                raise HTTPException(
                    400,
                    f"Matrix file must contain '{config.DURATION_SHEET}' "
                    f"and '{config.DISTANCE_SHEET}' sheets"
                )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Invalid matrix file: {e}")

        if dataset_id and save_to_dataset:
            await save_matrix_bytes(dataset_id, matrix_bytes)

        return Response(
            content    = matrix_bytes,
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers    = {"Content-Disposition": 'attachment; filename="matrix.xlsx"'},
        )

    # Build from OSRM
    coords, ids = [], []
    for dc_name, dc in config.DEPOTS.items():
        coords.append((dc["lat"], dc["lon"])); ids.append(dc_name)
    for s in stores_list:
        coords.append((s["lat"], s["lon"])); ids.append(s["node_id"])

    try:
        dist_mat, dur_mat = osrm_client.build_matrix_from_osrm(coords)
    except ConnectionError as e:
        raise HTTPException(503, str(e))

    import pandas as pd, io as _io
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame((dur_mat / 60).round(2), index=ids, columns=ids).to_excel(
            writer, sheet_name=config.DURATION_SHEET)
        pd.DataFrame(dist_mat, index=ids, columns=ids).to_excel(
            writer, sheet_name=config.DISTANCE_SHEET)
    buf.seek(0)
    matrix_bytes = buf.read()

    if dataset_id and save_to_dataset:
        await save_matrix_bytes(dataset_id, matrix_bytes)

    return Response(
        content    = matrix_bytes,
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": 'attachment; filename="matrix.xlsx"'},
    )


# ════════════════════════════════════════════════════════════
#  Run Groups
# ════════════════════════════════════════════════════════════

class RunGroupCreate(BaseModel):
    name      : str
    dataset_id: Optional[str] = None   # ← str now

class RunGroupRename(BaseModel):
    name: str


@app.get("/api/run-groups")
async def list_run_groups(db: AsyncIOMotorDatabase = Depends(get_db)):
    groups = await db[RunGroupDoc.COLLECTION].find().sort("created_at", -1).to_list(None)
    out    = []
    for g in groups:
        gd   = RunGroupDoc.to_dict(g)
        jobs = await db[JobDoc.COLLECTION].find(
            {"group_id": g["_id"]}).sort("created_at", 1).to_list(None)
        gd["jobs"] = []
        for j in jobs:
            jd  = JobDoc.to_dict(j)
            res = await db[JobResultDoc.COLLECTION].find_one({"_id": j["_id"]})
            if res:
                s = JobResultDoc.get_summary(res)
                jd["total_served"]    = s.get("total_served")    if s else None
                jd["total_unserved"]  = s.get("total_unserved")  if s else None
                jd["total_routes"]    = s.get("total_routes")    if s else None
                jd["total_cost"]      = s.get("total_cost")      if s else None
                jd["total_man_hours"] = s.get("total_man_hours") if s else None
            gd["jobs"].append(jd)
        out.append(gd)
    return out


@app.post("/api/run-groups")
async def create_run_group(body: RunGroupCreate,
                           db: AsyncIOMotorDatabase = Depends(get_db)):
    group_id = str(uuid.uuid4())
    doc      = RunGroupDoc.make(group_id, body.name, body.dataset_id)
    await db[RunGroupDoc.COLLECTION].insert_one(doc)
    return RunGroupDoc.to_dict(doc)


@app.patch("/api/run-groups/{group_id}")
async def rename_run_group(group_id: str, body: RunGroupRename,
                           db: AsyncIOMotorDatabase = Depends(get_db)):
    g = await db[RunGroupDoc.COLLECTION].find_one({"_id": group_id})
    if not g:
        raise HTTPException(404, "Group not found")
    await db[RunGroupDoc.COLLECTION].update_one(
        {"_id": group_id}, {"$set": {"name": body.name}})
    updated = await db[RunGroupDoc.COLLECTION].find_one({"_id": group_id})
    return RunGroupDoc.to_dict(updated)


@app.delete("/api/run-groups/{group_id}")
async def delete_run_group(group_id: str,
                           db: AsyncIOMotorDatabase = Depends(get_db)):
    g = await db[RunGroupDoc.COLLECTION].find_one({"_id": group_id})
    if not g:
        raise HTTPException(404, "Group not found")
    # Unlink jobs rather than delete them
    await db[JobDoc.COLLECTION].update_many(
        {"group_id": group_id}, {"$set": {"group_id": None}})
    await db[RunGroupDoc.COLLECTION].delete_one({"_id": group_id})
    return {"ok": True}


# ── Per-job version management ────────────────────────────────

class JobVersionPatch(BaseModel):
    version_name: Optional[str] = None
    group_id    : Optional[str] = None


@app.patch("/api/jobs/{job_id}/version")
async def patch_job_version(job_id: str, body: JobVersionPatch,
                            db: AsyncIOMotorDatabase = Depends(get_db)):
    job = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    updates = {}
    if body.version_name is not None: updates["version_name"] = body.version_name
    if body.group_id     is not None: updates["group_id"]     = body.group_id
    if updates:
        await db[JobDoc.COLLECTION].update_one({"_id": job_id}, {"$set": updates})
    updated = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    return JobDoc.to_dict(updated)


@app.post("/api/jobs/{job_id}/fork")
async def fork_job(job_id: str, db: AsyncIOMotorDatabase = Depends(get_db)):
    src = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not src:
        raise HTTPException(404, "Source job not found")
    src_res = await db[JobResultDoc.COLLECTION].find_one({"_id": job_id})
    if not src_res:
        raise HTTPException(404, "Source result not found")

    sibling_count = 0
    if src.get("group_id"):
        sibling_count = await db[JobDoc.COLLECTION].count_documents(
            {"group_id": src["group_id"]})

    new_id  = str(uuid.uuid4())
    new_doc = JobDoc.make(
        new_id,
        dataset_id   = src.get("dataset_id"),
        group_id     = src.get("group_id"),
        version_name = f"Manual v{sibling_count + 1}",
        is_manual    = True,
        mode         = src.get("mode"),
        max_trips    = src.get("max_trips"),
        solver_time  = src.get("solver_time"),
        season       = src.get("season"),
    )
    new_doc["status"]       = "done"
    new_doc["completed_at"] = datetime.datetime.utcnow()
    await db[JobDoc.COLLECTION].insert_one(new_doc)

    # Copy result (by value — edits don't affect original)
    new_res = {
        "_id"          : new_id,
        "summary_json" : src_res.get("summary_json"),
        "routes_json"  : src_res.get("routes_json"),
        "stops_json"   : src_res.get("stops_json"),
        "unserved_json": src_res.get("unserved_json"),
        "map_data_json": src_res.get("map_data_json"),
        "excel_file_id": None,  # heavy blob not copied — regen on next patch
    }
    await db[JobResultDoc.COLLECTION].insert_one(new_res)

    return {**JobDoc.to_dict(new_doc), "forked_from": job_id}


@app.patch("/api/jobs/{job_id}/result")
async def patch_job_result(job_id: str, body: dict,
                           db: AsyncIOMotorDatabase = Depends(get_db)):
    job = await db[JobDoc.COLLECTION].find_one({"_id": job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    res = await db[JobResultDoc.COLLECTION].find_one({"_id": job_id})
    if not res:
        raise HTTPException(404, "Job has no result to patch")

    updates = {}
    if "summary"       in body: updates["summary_json"]  = _dumps(body["summary"])
    if "route_summary" in body: updates["routes_json"]   = _dumps(body["route_summary"])
    if "stop_details"  in body: updates["stops_json"]    = _dumps(body["stop_details"])
    if "unserved"      in body: updates["unserved_json"] = _dumps(body["unserved"])
    if "map_data"      in body: updates["map_data_json"] = _dumps(body["map_data"])

    if updates:
        await db[JobResultDoc.COLLECTION].update_one({"_id": job_id}, {"$set": updates})

    # Recompute Excel if route/stop data changed
    if "route_summary" in body or "stop_details" in body:
        try:
            refreshed = await db[JobResultDoc.COLLECTION].find_one({"_id": job_id})
            excel = output_formatter.export_to_excel(
                body.get("route_summary", JobResultDoc.get_routes(refreshed)),
                body.get("stop_details",  JobResultDoc.get_stops(refreshed)),
                JobResultDoc.get_unserved(refreshed),
            )
            await save_excel_bytes(job_id, excel)
        except Exception as e:
            log.warning(f"Excel regen failed for {job_id}: {e}")

    return {"ok": True}


# ════════════════════════════════════════════════════════════
#  Frontend
# ════════════════════════════════════════════════════════════

@app.get("/")
def serve_frontend():
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return JSONResponse({"message": "VRP API v3 — see /docs"})

@app.websocket("/ws/logs/{job_id}")
async def ws_job_logs(ws: WebSocket, job_id: str):
    """
    Stream solver log lines to the browser in real time.

    Protocol:
      __PING__  → keep-alive, browser ignores
      __DONE__  → solve finished, browser closes connection
      anything else → a log line to display
    """
    await ws.accept()

    # Wait up to 8 s for the solver background task to register its queue
    for _ in range(80):
        if job_id in _job_log_queues:
            break
        await asyncio.sleep(0.1)

    q = _job_log_queues.get(job_id)
    if q is None:
        await ws.send_text("__DONE__")
        await ws.close()
        return

    try:
        while True:
            try:
                msg = await asyncio.wait_for(q.get(), timeout=60.0)
                await ws.send_text(msg)
                if msg == "__DONE__":
                    break
            except asyncio.TimeoutError:
                try:
                    await ws.send_text("__PING__")
                except Exception:
                    break
    except WebSocketDisconnect:
        pass
    finally:
        _job_log_queues.pop(job_id, None)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8067, reload=True)